import openpyxl
import requests


def get_urls_from_excel_list(file_path):
    urls = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    row_indices = [1 + x for x in range(0, ws.max_row)]
    for row_index in row_indices:
        url = ws['A' + str(row_index)].value
        urls.append(url)
    return urls


def get_page_content(urls):
    page_content = []
    for url in urls:
        r = requests.get(url)
        r.encoding = 'utf-8'  # all CK sites should use this.
        single_page_content = r.text
        page_content.append(single_page_content)
    return page_content
